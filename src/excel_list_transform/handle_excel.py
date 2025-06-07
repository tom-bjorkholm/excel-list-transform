#! /usr/bin/python3
"""Handle reading and writing of an excel file."""

# Copyright (c) 2024-2025 Tom Björkholm
# MIT License


# https://realpython.com/openpyxl-excel-spreadsheets-python/

import sys
from datetime import datetime, date, time
from copy import deepcopy
from typing import Optional, TypeAlias, Any
from decimal import Decimal
from xlsxwriter import Workbook as xlsxwriter_wb  # type: ignore
from pylightxl import readxl as pl_readxl  # type: ignore
from pylightxl import Database as pl_db, writexl as pl_writexl
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl import Workbook as openpyxl_wb
try:
    # only availabe in mypy, flake8 and pylint, not in python
    from openpyxl.cell import _CellGetValue  # pylint: disable=unused-import  # noqa: F401,E501
except ImportError:
    pass
from excel_list_transform.commontypes import NameData, NameDataMap, \
    NumData, NumDataSeq, NumRow
from excel_list_transform.config_enums import ExcelLib
from excel_list_transform.handle_empty_column import handle_empty_column_in_lst
from excel_list_transform.num_named_conversion import \
    named_cols_from_num_cols, num_cols_from_named_cols


_AnyCellValue: TypeAlias = Any  # AnyOf _CellGetValue # noqa: Y047
seentypes: list[str] = []


def fix_row_openpyxl(row: tuple[_AnyCellValue, ...],
                     max_column_reed: int) -> NumRow:
    """Convert types in row from openpyxl to those supported."""
    ret: NumRow = []
    colnum = 0
    for cell in row:
        if colnum >= max_column_reed:
            continue
        colnum += 1
        if cell is None:
            ret.append(None)
        elif isinstance(cell, (str, int, float, datetime)):
            assert isinstance(cell, (str, int, float, datetime))
            ret.append(cell)
        elif isinstance(cell, date):
            assert isinstance(cell, date)
            ret.append(datetime(year=cell.year,
                                month=cell.month,
                                day=cell.day))
        elif isinstance(cell, Decimal):
            assert isinstance(cell, Decimal)
            ret.append(float(cell))
        elif isinstance(cell, time):
            assert isinstance(cell, time)
            ret.append(cell.isoformat())
        else:
            typename = type(cell).__name__
            if typename not in seentypes:
                print('Sorry, cells with value of type ' +
                      f'{typename} are not supported\n' +
                      'Trying to convert to string.',
                      file=sys.stderr)
                seentypes.append(typename)
            ret.append(str(cell))
    return ret


def read_excel_openpyxl(filename: str,
                        max_column_read: int) \
                        -> NumData:
    """Read the excel file using openpyxl."""
    workbook = openpyxl_load_workbook(filename=filename, read_only=True,
                                      data_only=True)
    sheet = workbook.active
    assert sheet is not None
    res: NumData = []
    for row in sheet.iter_rows(values_only=True):
        res.append(fix_row_openpyxl(row, max_column_reed=max_column_read))
    return handle_empty_column_in_lst(res)


def read_excel_pyligthxl(filename: str,
                         max_column_read: int) \
                         -> NumData:
    """Read the excel file using pyligthxl."""
    pyl_db = pl_readxl(fn=filename)
    sheetnames = pyl_db.ws_names
    res = []
    for sheet in sheetnames:
        for row in pyl_db.ws(ws=sheet).rows:
            res.append(list(row)[:max_column_read])
    ret = []
    for row in res:
        row = [i if i != '' else None for i in row]
        ret.append(row)
    return ret


def read_excel_xlsxwriter(filename: str,
                          max_column_read: int) \
                          -> NumData:
    """Flag error and use other read library."""
    print('xlsxwriter cannot read, using pylightxl.', file=sys.stderr)
    return read_excel_pyligthxl(filename=filename,
                                max_column_read=max_column_read)


def excel_data_strip(data: NumData, strip_col_names: bool,
                     strip_values: bool) -> NumData:
    """Strip whitespace off titles and values as requested."""
    if not strip_values and not strip_col_names:
        return data
    datacopy: NumData = deepcopy(data)
    if strip_col_names:
        titlerow = datacopy[0]
        for index, titlecol in enumerate(titlerow):
            if isinstance(titlecol, str):
                titlerow[index] = titlecol.strip()
    if strip_values:
        for row in datacopy[1:]:
            for ind, col in enumerate(row):
                if isinstance(col, str):
                    row[ind] = col.strip()
    return datacopy


def read_excel_num(filename: str, max_column_read: int,
                   strip_col_names: bool, strip_values: bool,
                   excel_lib: Optional[ExcelLib] = None) \
                        -> NumData:
    """Read the excel file for number referenced columns."""
    dispatch = {ExcelLib.OPENPYXL: read_excel_openpyxl,
                ExcelLib.PYLIGHTXL: read_excel_pyligthxl,
                ExcelLib.XLSXWRITER: read_excel_xlsxwriter}
    if excel_lib is None:
        excel_lib = ExcelLib.PYLIGHTXL  # default for now
    data: NumData = dispatch[excel_lib](filename=filename,
                                        max_column_read=max_column_read)
    return excel_data_strip(data, strip_col_names=strip_col_names,
                            strip_values=strip_values)


def read_excel_named(filename: str, max_column_read: int,
                     strip_col_names: bool, strip_values: bool,
                     excel_lib: Optional[ExcelLib] = None) \
                        -> NameData:
    """Read the excel file for name referenced columns."""
    data = read_excel_num(filename=filename, max_column_read=max_column_read,
                          strip_col_names=strip_col_names,
                          strip_values=strip_values, excel_lib=excel_lib)
    return named_cols_from_num_cols(data=data, filename=filename)


def write_excel_openpyxl(data: NumData | NumDataSeq,
                         filename: str) -> None:
    """Write data as list in excel file using openpyxl."""
    workbook = openpyxl_wb()
    sheet = workbook.active
    assert sheet is not None
    workbook.encoding = 'utf-8'
    assert isinstance(data, list)
    for row in data:
        assert isinstance(row, (list, tuple))
        sheet.append(row)
    workbook.save(filename=filename)


def write_excel_pylightxl(data: NumData | NumDataSeq,
                          filename: str) -> None:
    """Write data as list in excel file using pylightxl."""
    pyl_db = pl_db()
    pyl_db.add_ws(ws="Data")
    for row_id, row in enumerate(data, start=1):
        for col_id, elem in enumerate(row, start=1):
            value = elem if elem is not None else ''
            pyl_db.ws(ws="Data").update_index(row=row_id, col=col_id,
                                              val=value)
    pl_writexl(db=pyl_db, fn=filename)


def write_excel_xlsxwriter(data: NumData | NumDataSeq,
                           filename: str) -> None:
    """Write data as list in excel file using xlsxwriter."""
    options = {'in_memory': True,
               'strings_to_numbers': False,
               'strings_to_formulaes': False,
               'strings_to_urls': False,
               'default_date_format': 'yyyy-mm-dd',
               'remove_timezone': True}
    with xlsxwriter_wb(filename=filename, options=options) as workbook:
        sheet = workbook.add_worksheet()
        assert isinstance(data, list)
        for i, row in enumerate(data):
            assert isinstance(row, (list, tuple))
            for j, elem in enumerate(row):
                if elem is not None:
                    if isinstance(elem, str):
                        sheet.write_string(row=i, col=j, string=elem)
                    elif isinstance(elem, int):
                        sheet.write_number(row=i, col=j, number=elem)
                    else:
                        msg = f'Unexpected data type {type(elem).__name__} '
                        msg += f'for row={i} col={j}. Value {elem}.'
                        print(msg, file=sys.stderr)
                        raise RuntimeError(msg)


def write_excel_num(data: NumData | NumDataSeq,
                    filename: str,
                    excel_lib: Optional[ExcelLib] = None) -> None:
    """Write data (number referenced columns) as list in excel file."""
    dispatch = {ExcelLib.OPENPYXL: write_excel_openpyxl,
                ExcelLib.PYLIGHTXL: write_excel_pylightxl,
                ExcelLib.XLSXWRITER: write_excel_xlsxwriter}
    if excel_lib is None:
        excel_lib = ExcelLib.XLSXWRITER  # default for now
    dispatch[excel_lib](data=data, filename=filename)


def write_excel_named(data: NameDataMap,
                      filename: str,
                      column_order: list[str],
                      excel_lib: Optional[ExcelLib] = None) -> None:
    """Write data (name referenced columns) as list in excel file."""
    wdata: NumData = num_cols_from_named_cols(data=data,
                                              column_order=column_order)
    write_excel_num(data=wdata, filename=filename,
                    excel_lib=excel_lib)
